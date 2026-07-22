#!/usr/bin/env python3
"""
apply-data.py — push the filled-in workbook back into the website.

Fill the yellow columns in Angel-Public-School-Data.xlsx, then:

    python3 apply-data.py              # dry run — shows what would change
    python3 apply-data.py --apply      # actually writes (creates .bak files)

Only rows where you filled the real-value column are touched. Everything else
is left exactly as it is and reported at the end so nothing goes silently
half-applied.

Requires openpyxl:  pip install openpyxl
"""

import argparse
import html
import re
import shutil
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is not installed.  Run:  pip install openpyxl")

ROOT = Path(__file__).resolve().parent
BOOK = ROOT / "Angel-Public-School-Data.xlsx"
PAGES = ["index.html", "about.html", "academics.html", "admissions.html",
         "contact.html", "404.html"]

BLANKS = {"", "(none)", "(none published)", "(not shown)", "(not set)",
          "(same as school hours)", "(already correct)", "(not mentioned)",
          "(mentioned in nav + footer)", "(mentioned in Our Focus, no tile)"}


def is_blank(v):
    return v is None or str(v).strip() in BLANKS


def esc(v):
    """Escape a spreadsheet value for insertion into HTML."""
    return html.escape(str(v).strip(), quote=False)


class Site:
    """Holds every page in memory; tracks edits until commit()."""

    def __init__(self):
        self.docs = {p: (ROOT / p).read_text(encoding="utf-8")
                     for p in PAGES if (ROOT / p).exists()}
        self.changes = []   # (page, what)
        self.skipped = []   # (sheet, what, why)

    def replace(self, old, new, label, pages=None):
        """Literal replacement across pages. Returns number of pages hit."""
        hits = 0
        for name in (pages or self.docs):
            doc = self.docs.get(name)
            if doc and old in doc:
                hits += 1              # the anchor text was found …
                if old == new:
                    continue           # … but the value is unchanged
                self.docs[name] = doc.replace(old, new)
                self.changes.append((name, f"{label}: {old!r} → {new!r}"))
        return hits

    def sub(self, pattern, repl, label, pages=None):
        """Regex replacement. Only counts as a change if the text actually differs."""
        hits = 0
        for name in (pages or self.docs):
            doc = self.docs.get(name)
            if doc is None:
                continue
            new_doc, n = re.subn(pattern, repl, doc)
            if n:
                hits += n
                if new_doc == doc:
                    continue           # matched, but produced identical output
                self.docs[name] = new_doc
                self.changes.append((name, f"{label} ({n}×)"))
        return hits

    def commit(self, apply):
        if not apply:
            return
        for name, text in self.docs.items():
            path = ROOT / name
            if path.read_text(encoding="utf-8") != text:
                shutil.copy2(path, path.with_suffix(path.suffix + ".bak"))
                path.write_text(text, encoding="utf-8")


# --------------------------------------------------------------------------
# sheet handlers
# --------------------------------------------------------------------------

def rows(ws, start=3):
    for r in ws.iter_rows(min_row=start, values_only=True):
        if any(c is not None and str(c).strip() for c in r):
            yield r


def do_profile(ws, site):
    """02 · School Profile — C = dummy on site, D = real value."""
    for r in rows(ws):
        _, field, dummy, real = r[0], r[1], r[2], r[3]
        if is_blank(real):
            continue
        if is_blank(dummy):
            site.skipped.append(("02 Profile", field,
                                 "no dummy text on the site to swap — needs a manual edit"))
            continue
        if not site.replace(esc(dummy), esc(real), f"profile/{field}"):
            site.skipped.append(("02 Profile", field,
                                 f"dummy text {dummy!r} not found in any page"))


def do_fees(ws, site):
    """03 · Fee Structure — rebuild the admissions.html fee table body."""
    body = []
    for r in rows(ws):
        cls, adm, tui, ann = r[0], r[1], r[2], r[3]
        if not cls or str(cls).strip() == "" or "concession" in str(cls).lower() \
           or "scholarship" in str(cls).lower():
            continue
        if any(v is None for v in (adm, tui, ann)):
            continue
        fmt = lambda v: f"₹ {v:,}" if isinstance(v, (int, float)) else f"₹ {v}"
        body.append(f"            <tr><td>{esc(cls)}</td><td>{fmt(adm)}</td>"
                    f"<td>{fmt(tui)}</td><td>{fmt(ann)}</td></tr>")

    if not body:
        site.skipped.append(("03 Fees", "fee table", "no complete rows found"))
        return

    pat = re.compile(
        r'(<th scope="col">Class</th>.*?<tbody>\n)(.*?)(\n\s*</tbody>)',
        re.S)
    site.sub(pat, lambda m: m.group(1) + "\n".join(body) + m.group(3),
             "fee table rebuilt", pages=["admissions.html"])


def do_dates(ws, site):
    """04 · Admission Dates — B = dummy date, C = real date."""
    for r in rows(ws):
        stage, dummy, real = r[0], r[1], r[2]
        if is_blank(real) or is_blank(dummy):
            continue
        if not site.replace(f"<td>{esc(dummy)}</td>", f"<td>{esc(real)}</td>",
                            f"date/{stage}", pages=["admissions.html"]):
            site.skipped.append(("04 Dates", stage, f"{dummy!r} not found in the table"))


def do_stats(ws, site):
    """05 · Statistics — target the number by its adjacent label, never blind-replace."""
    for r in rows(ws):
        _, label, dummy, real = r[0], r[1], r[2], r[3]
        if is_blank(real) or is_blank(dummy) or is_blank(label):
            continue
        hit = 0
        for kind in ("stat", "factbar"):
            pat = re.compile(
                r'(<span class="%s__num">)[^<]*(</span>\s*<span class="%s__label">%s</span>)'
                % (kind, kind, re.escape(esc(label))))
            hit += site.sub(pat, lambda m: m.group(1) + esc(real) + m.group(2),
                            f"stat/{label}")
        if not hit:
            site.skipped.append(("05 Stats", label, "no matching stat block found"))


def do_news(ws, site):
    """06 · News Items — D = real headline, E = real summary."""
    for r in rows(ws):
        slot, _cat, dummy_head, real_head, real_sum = r[0], r[1], r[2], r[3], r[4]
        if not is_blank(real_head) and not is_blank(dummy_head):
            if not site.replace(esc(dummy_head), esc(real_head), f"news/{slot}"):
                site.skipped.append(("06 News", slot,
                                     "headline not found — it may have been edited already"))
        if not is_blank(real_sum):
            site.skipped.append(("06 News", f"{slot} summary",
                                 "summaries must be pasted in by hand — excerpt text is not keyed"))


HANDLERS = {
    "02 · School Profile": do_profile,
    "03 · Fee Structure": do_fees,
    "04 · Admission Dates": do_dates,
    "05 · Statistics": do_stats,
    "06 · News Items": do_news,
}


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--apply", action="store_true",
                    help="write the changes (default is a dry run)")
    ap.add_argument("--book", default=str(BOOK), help="path to the workbook")
    args = ap.parse_args()

    book = Path(args.book)
    if not book.exists():
        sys.exit(f"Workbook not found: {book}")

    wb = load_workbook(book, data_only=True)
    site = Site()

    for name, fn in HANDLERS.items():
        if name in wb.sheetnames:
            fn(wb[name], site)
        else:
            print(f"  ! sheet missing from workbook: {name}")

    print(f"\n{'APPLIED' if args.apply else 'DRY RUN'} — {book.name}\n" + "=" * 68)

    if site.changes:
        print(f"\n{len(site.changes)} change(s):\n")
        by_page = {}
        for page, what in site.changes:
            by_page.setdefault(page, []).append(what)
        for page in sorted(by_page):
            print(f"  {page}")
            for what in by_page[page]:
                print(f"    · {what}")
    else:
        print("\nNothing to apply — no real-value cells were filled in.")

    if site.skipped:
        print(f"\n{len(site.skipped)} row(s) need a manual edit:\n")
        for sheet, what, why in site.skipped:
            print(f"  [{sheet}] {what}\n      {why}")

    site.commit(args.apply)

    if site.changes and not args.apply:
        print("\nRe-run with --apply to write these changes (.bak backups are created).")
    elif site.changes:
        print("\nWritten. Originals saved as *.html.bak")

    print()


if __name__ == "__main__":
    main()
